"""
Cierre de Curso e Historial Archivado.

Flujo:
  1. El administrador entra a `cierre_preview` (vista previa por jornada / curso entero).
  2. Confirma → `cierre_ejecutar` archiva todo en CierreCurso/MatriculaArchivada/AbonoArchivado
     y BORRA las matrículas + abonos del ciclo. Todo dentro de transaction.atomic.
  3. `cierre_historial` muestra todos los cierres pasados, agrupados por curso → jornada.
  4. `cierre_detalle` muestra el detalle de un cierre con TODOS los filtros
     (estado pago, modalidad, sede, búsqueda libre) y el mismo formato de tabla
     que la lista de matrículas viva.
  5. `cierre_export` descarga el detalle de un cierre como Excel profesional.
"""

from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    Abono, AbonoArchivado, CierreAdministrativo, CierreCurso, Curso, Estudiante,
    EstudianteArchivado, JornadaCurso, Matricula, MatriculaArchivada,
)
from .permisos import admin_requerido, matricula_requerida


MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ═════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════

def _matriculas_a_cerrar(curso, jornada=None):
    """
    Devuelve la queryset de matrículas que se incluirán en el cierre.
    - Si jornada se pasa: solo esa jornada.
    - Si no: TODAS las matrículas del curso.
    """
    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    ).filter(curso=curso)
    if jornada is not None:
        qs = qs.filter(jornada=jornada)
    return qs.order_by('estudiante__apellidos', 'estudiante__nombres')


def _calcular_totales(matriculas):
    """Calcula totales de un conjunto de matrículas (vivas o archivadas)."""
    total_facturado = Decimal('0.00')
    total_cobrado = Decimal('0.00')
    total_pendiente = Decimal('0.00')
    c_pagado = c_parcial = c_pendiente = c_retiro = 0

    for m in matriculas:
        # Soporta tanto Matricula viva como MatriculaArchivada
        if hasattr(m, 'valor_neto') and not callable(m.valor_neto):
            vn = m.valor_neto if isinstance(m.valor_neto, Decimal) else Decimal(str(m.valor_neto))
        else:
            vc = m.valor_curso or Decimal('0.00')
            desc = getattr(m, 'descuento', None) or Decimal('0.00')
            vn = max(Decimal('0.00'), vc - desc)
        vp = m.valor_pagado or Decimal('0.00')

        total_facturado += vn
        total_cobrado += vp

        estado_pago = getattr(m, 'estado_pago', None)
        if callable(estado_pago):
            estado_pago = estado_pago()
        if not estado_pago:
            # Calcular manualmente
            estado = getattr(m, 'estado', '')
            if estado == 'retiro_voluntario':
                estado_pago = 'Retiro'
            elif vp >= vn and vn > 0:
                estado_pago = 'Pagado'
            elif vp > 0:
                estado_pago = 'Parcial'
            else:
                estado_pago = 'Pendiente'

        if estado_pago == 'Retiro':
            c_retiro += 1
        elif estado_pago == 'Pagado':
            c_pagado += 1
            # Pagados no aportan saldo pendiente
        elif estado_pago == 'Parcial':
            c_parcial += 1
            total_pendiente += max(Decimal('0.00'), vn - vp)
        else:
            c_pendiente += 1
            total_pendiente += max(Decimal('0.00'), vn - vp)

    return {
        'total_matriculas': len(matriculas) if hasattr(matriculas, '__len__') else matriculas.count(),
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'total_pendiente': total_pendiente,
        'conteo_pagado': c_pagado,
        'conteo_parcial': c_parcial,
        'conteo_pendiente': c_pendiente,
        'conteo_retiro': c_retiro,
    }


def _snapshot_matricula(matricula, cierre):
    """Crea una MatriculaArchivada a partir de una Matricula viva."""
    est = matricula.estudiante
    curso = matricula.curso
    jornada = matricula.jornada

    archivada = MatriculaArchivada.objects.create(
        cierre=cierre,
        matricula_original_id=matricula.pk,

        # FKs débiles
        estudiante=est,
        curso=curso,
        jornada=jornada,

        # Estudiante (denormalizado)
        cedula=est.cedula if est else '',
        apellidos=est.apellidos if est else '',
        nombres=est.nombres if est else '',
        edad=str(est.edad) if (est and est.edad is not None) else '',
        correo=est.correo if est else '',
        celular=est.celular if est else '',
        ciudad_estudiante=est.ciudad if est else '',
        nivel_formacion=est.get_nivel_formacion_display() if (est and est.nivel_formacion) else '',
        talla_camiseta=matricula.talla_camiseta or '',

        # Curso / Jornada (denormalizado)
        curso_nombre=curso.nombre if curso else '',
        curso_categoria=curso.categoria.nombre if (curso and curso.categoria) else '',
        jornada_descripcion=(jornada.descripcion_legible if jornada else ''),
        jornada_fecha_inicio=(jornada.fecha_inicio if jornada else None),
        jornada_horario=matricula.horario if matricula.horario != '—' else '',
        sede=matricula.sede if matricula.sede != '—' else '',

        # Matrícula
        modalidad=matricula.modalidad,
        tipo_matricula=matricula.tipo_matricula or '',
        estado=matricula.estado or '',
        fecha_matricula=matricula.fecha_matricula,
        valor_curso=matricula.valor_curso or Decimal('0.00'),
        descuento=matricula.descuento or Decimal('0.00'),
        valor_neto=matricula.valor_neto,
        valor_pagado=matricula.valor_pagado or Decimal('0.00'),
        saldo=matricula.saldo,
        estado_pago=matricula.estado_pago,

        # Comprobante
        tipo_registro=matricula.tipo_registro or '',
        factura_realizada=matricula.factura_realizada or '',
        fact_nombres=matricula.fact_nombres or '',
        fact_apellidos=matricula.fact_apellidos or '',
        fact_cedula=matricula.fact_cedula or '',
        fact_correo=matricula.fact_correo or '',
        link_comprobante=matricula.link_comprobante or '',

        observaciones=matricula.observaciones or '',
        registrado_por_nombre=(
            matricula.registrado_por.get_full_name() or matricula.registrado_por.username
            if matricula.registrado_por_id else ''
        ),
        creado_original=matricula.creado,
    )

    # Archivar abonos
    for abono in matricula.abonos.all():
        AbonoArchivado.objects.create(
            matricula_archivada=archivada,
            cierre=cierre,
            abono_original_id=abono.pk,
            fecha=abono.fecha,
            monto=abono.monto,
            tipo_pago=abono.tipo_pago or '',
            tipo_pago_label=abono.get_tipo_pago_display() if abono.tipo_pago else '',
            numero_modulo=abono.numero_modulo,
            cuenta_para_saldo=abono.cuenta_para_saldo,
            metodo=abono.metodo or '',
            metodo_label=abono.get_metodo_display() if abono.metodo else '',
            banco=abono.banco or '',
            banco_label=abono.get_banco_display() if abono.banco else '',
            numero_recibo=abono.numero_recibo or '',
            observaciones=abono.observaciones or '',
            registrado_por_nombre=(
                abono.registrado_por.get_full_name() or abono.registrado_por.username
                if abono.registrado_por_id else ''
            ),
            creado_original=abono.creado,
        )

    return archivada


def _snapshot_estudiante(estudiante, cierre):
    """Crea un EstudianteArchivado a partir de un Estudiante vivo (snapshot)."""
    return EstudianteArchivado.objects.create(
        cierre=cierre,
        estudiante_original_id=estudiante.pk,
        cedula=estudiante.cedula,
        apellidos=estudiante.apellidos,
        nombres=estudiante.nombres,
        edad=estudiante.edad,
        correo=estudiante.correo or '',
        celular=estudiante.celular or '',
        nivel_formacion=(
            estudiante.get_nivel_formacion_display() if estudiante.nivel_formacion else ''
        ),
        titulo_profesional=estudiante.titulo_profesional or '',
        ciudad=estudiante.ciudad or '',
        creado_original=estudiante.creado,
    )


# ═════════════════════════════════════════════════════════════════
# Vista previa del cierre
# ═════════════════════════════════════════════════════════════════

@admin_requerido
def cierre_preview(request, curso_pk):
    """
    Muestra qué matrículas y abonos se archivarán.
    El admin elige: cerrar UNA jornada específica, o TODO el curso.
    """
    curso = get_object_or_404(Curso, pk=curso_pk)

    jornada_id = request.GET.get('jornada', '').strip()
    jornada = None
    if jornada_id.isdigit():
        jornada = get_object_or_404(JornadaCurso, pk=int(jornada_id), curso=curso)

    matriculas = list(_matriculas_a_cerrar(curso, jornada))
    totales = _calcular_totales(matriculas)
    total_abonos = sum(m.abonos.count() for m in matriculas)

    # Todas las jornadas del curso (para que pueda escoger otra desde la vista previa)
    jornadas_todas = curso.jornadas.all().order_by('modalidad', 'fecha_inicio')

    return render(request, 'cursos/cierre_preview.html', {
        'curso': curso,
        'jornada': jornada,
        'jornadas_todas': jornadas_todas,
        'matriculas': matriculas,
        'totales': totales,
        'total_abonos': total_abonos,
        'alcance': 'jornada' if jornada else 'curso',
    })


@admin_requerido
@require_POST
def cierre_ejecutar(request, curso_pk):
    """
    Ejecuta el cierre: archiva todo y borra las matrículas/abonos vivos.
    Operación atómica: si algo falla, no se borra nada.

    Si llega `limpiar_directorio=on`, además archiva y borra los estudiantes
    que queden huérfanos (sin matrículas vivas en otros cursos).
    """
    curso = get_object_or_404(Curso, pk=curso_pk)

    jornada_id = request.POST.get('jornada_id', '').strip()
    ciclo_etiqueta = (request.POST.get('ciclo_etiqueta', '').strip())[:80]
    observaciones = request.POST.get('observaciones', '').strip()
    limpiar_directorio = request.POST.get('limpiar_directorio') == 'on'

    jornada = None
    if jornada_id.isdigit():
        jornada = get_object_or_404(JornadaCurso, pk=int(jornada_id), curso=curso)

    qs = _matriculas_a_cerrar(curso, jornada)
    matriculas = list(qs)

    if not matriculas:
        messages.warning(
            request,
            'No hay matrículas para cerrar en ese alcance. No se hizo ningún cambio.'
        )
        return redirect('academia:cursos_lista', modalidad='presencial')

    totales = _calcular_totales(matriculas)

    # ── Identificar estudiantes que podrían quedar huérfanos ──
    # (los que solo tienen matrículas en este alcance del cierre)
    estudiantes_a_evaluar_ids = set(m.estudiante_id for m in matriculas if m.estudiante_id)

    try:
        with transaction.atomic():
            cierre = CierreCurso.objects.create(
                curso=curso,
                curso_nombre=curso.nombre,
                curso_categoria=(curso.categoria.nombre if curso.categoria else ''),
                jornada=jornada,
                jornada_descripcion=(jornada.descripcion_legible if jornada else ''),
                jornada_modalidad=(jornada.modalidad if jornada else ''),
                jornada_fecha_inicio=(jornada.fecha_inicio if jornada else None),
                jornada_sede=(jornada.ciudad if jornada else ''),
                alcance='jornada' if jornada else 'curso',
                ciclo_etiqueta=ciclo_etiqueta,
                observaciones=observaciones,
                cerrado_por=request.user if request.user.is_authenticated else None,
                **{k: v for k, v in totales.items() if k != 'total_matriculas'},
                total_matriculas=totales['total_matriculas'],
            )

            ids_a_borrar = []
            for m in matriculas:
                _snapshot_matricula(m, cierre)
                ids_a_borrar.append(m.pk)

            Matricula.objects.filter(pk__in=ids_a_borrar).delete()

            # ── Limpieza opcional del directorio de estudiantes ──
            estudiantes_archivados = 0
            if limpiar_directorio and estudiantes_a_evaluar_ids:
                # Tras borrar las matrículas, ¿cuáles estudiantes quedaron sin matrículas Y sin adicionales?
                # No tocamos estudiantes con Adicional protegido (exámenes supletorios cobrados, etc.)
                huerfanos = Estudiante.objects.filter(
                    pk__in=estudiantes_a_evaluar_ids,
                    matriculas__isnull=True,
                    adicionales__isnull=True,
                ).distinct()

                for est in huerfanos:
                    _snapshot_estudiante(est, cierre)
                    estudiantes_archivados += 1

                # Borrar los huérfanos del directorio vivo
                huerfanos.delete()

            cierre.limpio_directorio = limpiar_directorio
            cierre.total_estudiantes_archivados = estudiantes_archivados
            cierre.save(update_fields=['limpio_directorio', 'total_estudiantes_archivados'])

    except Exception as e:
        messages.error(
            request,
            f'Ocurrió un error durante el cierre. No se modificó nada. Detalle: {e}'
        )
        return redirect('academia:cursos_lista', modalidad='presencial')

    extra_msg = ''
    if limpiar_directorio:
        if estudiantes_archivados:
            extra_msg = (
                f' Además se archivaron y limpiaron {estudiantes_archivados} '
                f'estudiante(s) que quedaron sin matrículas activas.'
            )
        else:
            extra_msg = (
                ' No hubo estudiantes huérfanos para archivar '
                '(todos siguen matriculados en otros cursos).'
            )

    if jornada:
        messages.success(
            request,
            f'✅ Cierre realizado. Se archivaron {totales["total_matriculas"]} matrícula(s) '
            f'de la jornada "{jornada.descripcion_legible}" del curso "{curso.nombre}".{extra_msg} '
            f'Puedes consultarlas en Historial de cursos cerrados.'
        )
    else:
        messages.success(
            request,
            f'✅ Cierre realizado. Se archivaron {totales["total_matriculas"]} matrícula(s) '
            f'de TODAS las jornadas del curso "{curso.nombre}".{extra_msg} '
            f'Puedes consultarlas en Historial de cursos cerrados.'
        )

    return redirect('academia:cierre_detalle', cierre_pk=cierre.pk)


# ═════════════════════════════════════════════════════════════════
# CIERRE GLOBAL — todos los cursos de una modalidad (o ambos)
# ═════════════════════════════════════════════════════════════════

@admin_requerido
def cierre_global_preview(request, modalidad):
    """
    Vista previa del cierre global: muestra TODAS las matrículas de la
    modalidad indicada ('presencial', 'online' o 'todas') agrupadas por curso.
    """
    if modalidad not in ('presencial', 'online', 'todas'):
        messages.error(request, 'Modalidad no válida para cierre global.')
        return redirect('academia:cursos_lista', modalidad='presencial')

    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    )
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    qs = qs.order_by('curso__nombre', 'estudiante__apellidos', 'estudiante__nombres')

    matriculas = list(qs)
    totales = _calcular_totales(matriculas)
    total_abonos = sum(m.abonos.count() for m in matriculas)

    # Agrupar por curso para vista previa
    grupos = defaultdict(list)
    for m in matriculas:
        grupos[(m.curso_id, m.curso.nombre if m.curso else 'Sin curso')].append(m)
    grupos_lista = sorted(
        [{'curso_id': k[0], 'curso_nombre': k[1], 'matriculas': v, 'total': len(v),
          'subtotal_facturado': sum(mm.valor_neto for mm in v),
          'subtotal_cobrado': sum(mm.valor_pagado or Decimal('0.00') for mm in v)}
         for k, v in grupos.items()],
        key=lambda g: g['curso_nombre'].lower()
    )

    cursos_distintos = len(grupos_lista)

    label_modalidad = {
        'presencial': '🏫 Presencial',
        'online': '🟢 Online',
        'todas': '🌐 TODAS las modalidades',
    }[modalidad]

    return render(request, 'cursos/cierre_global_preview.html', {
        'modalidad': modalidad,
        'modalidad_label': label_modalidad,
        'matriculas': matriculas,
        'grupos': grupos_lista,
        'cursos_distintos': cursos_distintos,
        'totales': totales,
        'total_abonos': total_abonos,
    })


@admin_requerido
@require_POST
def cierre_global_ejecutar(request, modalidad):
    """
    Ejecuta un cierre GLOBAL: archiva TODAS las matrículas de la modalidad
    (o de ambas si modalidad='todas') y borra todo, creando UN único
    CierreCurso por cada curso afectado (para poder filtrar/consultar luego
    de forma granular). Y opcionalmente limpia el directorio de estudiantes.
    """
    if modalidad not in ('presencial', 'online', 'todas'):
        messages.error(request, 'Modalidad no válida.')
        return redirect('academia:cursos_lista', modalidad='presencial')

    ciclo_etiqueta = (request.POST.get('ciclo_etiqueta', '').strip())[:80]
    observaciones = request.POST.get('observaciones', '').strip()
    limpiar_directorio = request.POST.get('limpiar_directorio') == 'on'

    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    )
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)

    matriculas_all = list(qs)
    if not matriculas_all:
        messages.warning(request, 'No hay matrículas que cerrar en ese alcance.')
        return redirect('academia:cursos_lista', modalidad=modalidad if modalidad != 'todas' else 'presencial')

    # Identificar estudiantes a evaluar
    estudiantes_a_evaluar_ids = set(m.estudiante_id for m in matriculas_all if m.estudiante_id)

    # Agrupar por curso (para crear un cierre por curso, no un cierre gigante único)
    por_curso = defaultdict(list)
    for m in matriculas_all:
        por_curso[m.curso_id].append(m)

    cierres_creados = []
    total_archivado = 0

    try:
        with transaction.atomic():
            for curso_id, matriculas_curso in por_curso.items():
                curso = matriculas_curso[0].curso  # mismo en todas
                totales = _calcular_totales(matriculas_curso)

                cierre = CierreCurso.objects.create(
                    curso=curso,
                    curso_nombre=curso.nombre if curso else 'Sin curso',
                    curso_categoria=(curso.categoria.nombre if curso and curso.categoria else ''),
                    jornada=None,
                    jornada_descripcion='',
                    jornada_modalidad='',
                    alcance='global',
                    modalidad_global=modalidad,
                    ciclo_etiqueta=ciclo_etiqueta or f'Cierre global {modalidad}',
                    observaciones=observaciones,
                    cerrado_por=request.user if request.user.is_authenticated else None,
                    total_matriculas=totales['total_matriculas'],
                    total_facturado=totales['total_facturado'],
                    total_cobrado=totales['total_cobrado'],
                    total_pendiente=totales['total_pendiente'],
                    conteo_pagado=totales['conteo_pagado'],
                    conteo_parcial=totales['conteo_parcial'],
                    conteo_pendiente=totales['conteo_pendiente'],
                    conteo_retiro=totales['conteo_retiro'],
                )
                cierres_creados.append(cierre)
                total_archivado += len(matriculas_curso)

                ids = []
                for m in matriculas_curso:
                    _snapshot_matricula(m, cierre)
                    ids.append(m.pk)
                Matricula.objects.filter(pk__in=ids).delete()

            # ── Limpieza del directorio ──
            estudiantes_archivados_total = 0
            if limpiar_directorio and estudiantes_a_evaluar_ids:
                huerfanos = Estudiante.objects.filter(
                    pk__in=estudiantes_a_evaluar_ids,
                    matriculas__isnull=True,
                    adicionales__isnull=True,
                ).distinct()
                # Asignamos los snapshots al primer cierre creado (a modo de "ancla")
                cierre_ancla = cierres_creados[0]
                for est in huerfanos:
                    _snapshot_estudiante(est, cierre_ancla)
                    estudiantes_archivados_total += 1
                huerfanos.delete()

                # Marcar todos los cierres del global
                for c in cierres_creados:
                    c.limpio_directorio = True
                    c.save(update_fields=['limpio_directorio'])
                cierre_ancla.total_estudiantes_archivados = estudiantes_archivados_total
                cierre_ancla.save(update_fields=['total_estudiantes_archivados'])

    except Exception as e:
        messages.error(
            request,
            f'Ocurrió un error durante el cierre global. No se modificó nada. Detalle: {e}'
        )
        return redirect('academia:cursos_lista', modalidad=modalidad if modalidad != 'todas' else 'presencial')

    extra = ''
    if limpiar_directorio:
        extra = (
            f' Y se archivaron {estudiantes_archivados_total} estudiante(s) huérfano(s) del directorio.'
            if estudiantes_archivados_total
            else ' (No hubo estudiantes huérfanos para archivar.)'
        )

    messages.success(
        request,
        f'✅ Cierre global ejecutado. Se archivaron {total_archivado} matrícula(s) '
        f'en {len(cierres_creados)} cierre(s) (uno por curso afectado).{extra} '
        f'Puedes consultarlos en Historial de cursos cerrados.'
    )
    return redirect('academia:cierre_historial')


# ═════════════════════════════════════════════════════════════════
# Historial de cierres
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def archivo_index(request):
    """
    Página índice del Archivo, organizada en CARPETAS → SUBCARPETAS por MES.

    Estructura:
      📁 Estudiantes
         └─ 📅 Mayo 2026   (5 cierre(s) · 9 matrícula(s) · 6 estudiante(s))
         └─ 📅 Junio 2026
         └─ 📅 Julio 2026
         ...
      📁 Administrativo
         └─ 📅 Mayo 2026   (1 corte(s))
         └─ 📅 Junio 2026
         ...

    Cada subcarpeta-mes enlaza a la lista correspondiente filtrada por ?anio=YYYY&mes=MM,
    para que NUNCA se mezclen registros de meses distintos en una sola pantalla.
    """
    es_admin_user = request.user.is_superuser or request.user.groups.filter(
        name='Administradores'
    ).exists()

    # ───────────────────────────────────────────────────────────────
    # CARPETA ESTUDIANTES — agrupar cierres de curso + estudiantes
    # archivados por (año, mes) basándose en fecha_cierre / archivado_en
    # ───────────────────────────────────────────────────────────────
    # Recorremos los cierres una sola vez y construimos un diccionario:
    #   periodos_est[(anio, mes)] = {
    #       'cierres': N,
    #       'matriculas': N,
    #       'estudiantes_arch': N,
    #   }
    periodos_est = defaultdict(lambda: {
        'cierres': 0,
        'matriculas': 0,
        'estudiantes_arch': 0,
    })

    # Cierres de curso (cada uno trae sus propios totales congelados)
    for c in CierreCurso.objects.all().only(
        'id', 'fecha_cierre', 'total_matriculas'
    ):
        key = (c.fecha_cierre.year, c.fecha_cierre.month)
        periodos_est[key]['cierres'] += 1
        periodos_est[key]['matriculas'] += c.total_matriculas

    # Estudiantes archivados (por su propia fecha de archivado)
    for ea in EstudianteArchivado.objects.all().only('archivado_en'):
        key = (ea.archivado_en.year, ea.archivado_en.month)
        periodos_est[key]['estudiantes_arch'] += 1

    # Convertimos el diccionario en una lista ordenada de períodos
    # (más reciente primero) listos para el template.
    meses_estudiantes = []
    for (anio, mes), datos in sorted(periodos_est.items(), reverse=True):
        meses_estudiantes.append({
            'anio': anio,
            'mes': mes,
            'etiqueta': f'{MESES_ES[mes]} {anio}',
            'total_cierres': datos['cierres'],
            'total_matriculas': datos['matriculas'],
            'total_estudiantes_arch': datos['estudiantes_arch'],
            # Descripción corta para mostrar bajo el nombre del mes
            'descripcion': (
                f'{datos["cierres"]} cierre(s) · '
                f'{datos["matriculas"]} matrícula(s) · '
                f'{datos["estudiantes_arch"]} estudiante(s)'
            ),
        })

    total_cierres_curso = CierreCurso.objects.count()
    total_estudiantes_arch = EstudianteArchivado.objects.count()

    # ───────────────────────────────────────────────────────────────
    # CARPETA ADMINISTRATIVO — agrupar cierres administrativos por mes
    # ───────────────────────────────────────────────────────────────
    meses_admin = []
    if es_admin_user:
        periodos_adm = defaultdict(int)
        for ca in CierreAdministrativo.objects.all().only('anio', 'mes', 'fecha_cierre'):
            # Si el cierre tiene `mes` explícito, lo usamos (ya está modelado por mes).
            # Si no (cierre anual completo), caemos a la fecha_cierre real.
            anio = ca.anio
            mes = ca.mes if ca.mes else ca.fecha_cierre.month
            periodos_adm[(anio, mes)] += 1

        for (anio, mes), n in sorted(periodos_adm.items(), reverse=True):
            meses_admin.append({
                'anio': anio,
                'mes': mes,
                'etiqueta': f'{MESES_ES[mes]} {anio}',
                'total_cierres': n,
                'descripcion': f'{n} corte(s) de caja',
            })

    total_cierres_admin = CierreAdministrativo.objects.count()

    # ───────────────────────────────────────────────────────────────
    # Carpetas finales para el template
    # ───────────────────────────────────────────────────────────────
    carpetas = [
        {
            'clave': 'estudiantes',
            'titulo': 'Estudiantes',
            'icono': '🎓',
            'descripcion': 'Cursos cerrados y estudiantes archivados, organizados por mes.',
            'color': '#1a237e',
            'meses': meses_estudiantes,
            # Cada mes tiene DOS sub-enlaces (cursos cerrados / estudiantes archivados);
            # los nombres de URL se resuelven en el template.
            'url_cursos': 'academia:cierre_historial',
            'url_estudiantes': 'academia:estudiantes_archivados_lista',
            'resumen': f'{total_cierres_curso + total_estudiantes_arch} registro(s) en {len(meses_estudiantes)} mes(es)',
            'vacio_msg': 'Aún no hay cursos cerrados ni estudiantes archivados.',
            'visible': True,
        },
    ]

    if es_admin_user:
        carpetas.append({
            'clave': 'administrativo',
            'titulo': 'Administrativo',
            'icono': '💰',
            'descripcion': 'Cortes de caja y cierres financieros, organizados por mes (solo administradores).',
            'color': '#2e7d32',
            'meses': meses_admin,
            'url_admin': 'academia:cierre_admin_historial',
            'resumen': f'{total_cierres_admin} corte(s) en {len(meses_admin)} mes(es)',
            'vacio_msg': 'Aún no hay cortes de caja archivados.',
            'visible': True,
        })

    return render(request, 'historial/archivo_index.html', {
        'carpetas': carpetas,
        'es_admin_user': es_admin_user,
    })


@matricula_requerida
def cierre_historial(request):
    """
    Lista todos los cierres agrupados por CURSO → cierres (jornadas / curso entero).
    Permite filtrar por curso, modalidad, año.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    q = request.GET.get('q', '').strip()

    qs = CierreCurso.objects.all().order_by('-fecha_cierre')

    if curso_id.isdigit():
        qs = qs.filter(curso_id=int(curso_id))
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(jornada_modalidad=modalidad)
    if anio.isdigit():
        qs = qs.filter(fecha_cierre__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha_cierre__month=int(mes))
    if q:
        qs = qs.filter(
            Q(curso_nombre__icontains=q)
            | Q(jornada_descripcion__icontains=q)
            | Q(ciclo_etiqueta__icontains=q)
            | Q(jornada_sede__icontains=q)
        )

    # Agrupar por curso (string para tolerar cursos eliminados)
    grupos_dict = defaultdict(list)
    for cierre in qs:
        clave = (cierre.curso_id, cierre.curso_nombre)
        grupos_dict[clave].append(cierre)

    grupos = []
    for (curso_id_k, nombre), cierres in sorted(grupos_dict.items(), key=lambda x: x[0][1].lower()):
        total_mat = sum(c.total_matriculas for c in cierres)
        total_fact = sum(c.total_facturado for c in cierres)
        total_cob = sum(c.total_cobrado for c in cierres)
        grupos.append({
            'curso_id': curso_id_k,
            'curso_nombre': nombre,
            'cierres': cierres,
            'total_cierres': len(cierres),
            'total_matriculas': total_mat,
            'total_facturado': total_fact,
            'total_cobrado': total_cob,
        })

    cursos = Curso.objects.all().order_by('nombre')
    anios = sorted(
        set(CierreCurso.objects.dates('fecha_cierre', 'year').values_list('fecha_cierre__year', flat=True)),
        reverse=True
    )

    return render(request, 'historial/cierres_lista.html', {
        'grupos': grupos,
        'cursos': cursos,
        'anios': anios,
        'total_cierres': qs.count(),
        'filtros': {
            'curso': curso_id,
            'modalidad': modalidad,
            'anio': anio,
            'mes': mes,
            'q': q,
        },
    })


@matricula_requerida
def cierre_detalle(request, cierre_pk):
    """
    Detalle de un cierre: tabla completa con los mismos filtros que la lista de matrículas
    viva (Pagado/Parcial/Pendiente, modalidad, sede, búsqueda libre).
    """
    cierre = get_object_or_404(CierreCurso, pk=cierre_pk)

    # Queryset base con prefetch de abonos archivados
    matriculas_qs = cierre.matriculas_archivadas.prefetch_related('abonos_archivados').all()

    # ── Filtros ──
    estado_pago = request.GET.get('estado', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    sede = request.GET.get('sede', '').strip()
    q = request.GET.get('q', '').strip()

    if estado_pago in ('Pagado', 'Parcial', 'Pendiente', 'Retiro'):
        matriculas_qs = matriculas_qs.filter(estado_pago=estado_pago)
    if modalidad in ('presencial', 'online'):
        matriculas_qs = matriculas_qs.filter(modalidad=modalidad)
    if sede:
        matriculas_qs = matriculas_qs.filter(sede__iexact=sede)
    if q:
        matriculas_qs = matriculas_qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
            | Q(correo__icontains=q)
            | Q(celular__icontains=q)
            | Q(jornada_descripcion__icontains=q)
        )

    matriculas = list(matriculas_qs.order_by('apellidos', 'nombres'))

    # Sedes distintas del cierre (para el selector)
    sedes_disponibles = sorted(
        set(s for s in cierre.matriculas_archivadas.values_list('sede', flat=True) if s)
    )

    # Recalcular conteo por estado para los chips superiores
    conteo_filtrado = {'Pagado': 0, 'Parcial': 0, 'Pendiente': 0, 'Retiro': 0}
    for m in matriculas:
        conteo_filtrado[m.estado_pago] = conteo_filtrado.get(m.estado_pago, 0) + 1

    totales_filtrados = _calcular_totales(matriculas)

    return render(request, 'historial/cierre_detalle.html', {
        'cierre': cierre,
        'matriculas': matriculas,
        'sedes_disponibles': sedes_disponibles,
        'conteo_filtrado': conteo_filtrado,
        'totales_filtrados': totales_filtrados,
        'filtros': {
            'estado': estado_pago,
            'modalidad': modalidad,
            'sede': sede,
            'q': q,
        },
    })


@admin_requerido
@require_POST
def cierre_eliminar(request, cierre_pk):
    """Elimina un cierre del historial (los archivados se borran en cascada)."""
    cierre = get_object_or_404(CierreCurso, pk=cierre_pk)
    nombre = cierre.encabezado
    cierre.delete()
    messages.success(request, f'Cierre eliminado: {nombre}.')
    return redirect('academia:cierre_historial')


# ═════════════════════════════════════════════════════════════════
# Exportación a Excel del detalle de un cierre
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def cierre_export(request, cierre_pk):
    """Descarga el cierre completo como Excel: una hoja con matrículas + una con abonos."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cierre = get_object_or_404(CierreCurso, pk=cierre_pk)

    # Aplicar mismos filtros del detalle si vienen por GET (para exportar el filtro vivo)
    matriculas_qs = cierre.matriculas_archivadas.prefetch_related('abonos_archivados').all()
    estado_pago = request.GET.get('estado', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    sede = request.GET.get('sede', '').strip()
    q = request.GET.get('q', '').strip()

    if estado_pago in ('Pagado', 'Parcial', 'Pendiente', 'Retiro'):
        matriculas_qs = matriculas_qs.filter(estado_pago=estado_pago)
    if modalidad in ('presencial', 'online'):
        matriculas_qs = matriculas_qs.filter(modalidad=modalidad)
    if sede:
        matriculas_qs = matriculas_qs.filter(sede__iexact=sede)
    if q:
        matriculas_qs = matriculas_qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    matriculas = list(matriculas_qs.order_by('apellidos', 'nombres'))

    wb = Workbook()

    # ── Estilos ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(bold=True, size=14, color='1A237E')
    subtitle_font = Font(italic=True, size=10, color='555555')
    money_font = Font(color='2E7D32', bold=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_fill = PatternFill('solid', fgColor='FFF8E1')

    # ── Hoja 1: Matrículas ──
    ws = wb.active
    ws.title = 'Matrículas'

    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Edad',
        'Correo', 'Celular', 'Ciudad', 'Nivel',
        'Curso', 'Categoría', 'Jornada', 'Sede', 'Modalidad', 'Horario',
        'Tipo matrícula', 'Estado matrícula',
        'Fecha matrícula', 'Valor curso', 'Descuento', 'Valor neto',
        'Pagado', 'Saldo', 'Estado pago',
        'Talla', 'Tipo registro', 'Factura', 'Observaciones', 'Registrado por',
    ]

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=f'Historial archivado — {cierre.encabezado}').font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subt = ws.cell(
        row=2, column=1,
        value=f'Cerrado el {cierre.fecha_cierre.strftime("%d/%m/%Y %H:%M")} · '
              f'{len(matriculas)} matrícula(s) en este filtro · '
              f'Alcance: {cierre.get_alcance_display()}'
    )
    subt.font = subtitle_font
    subt.alignment = Alignment(horizontal='center')

    # Encabezados
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    # Datos
    for row_idx, m in enumerate(matriculas, start=5):
        values = [
            m.cedula, m.apellidos, m.nombres, m.edad,
            m.correo, m.celular, m.ciudad_estudiante, m.nivel_formacion,
            m.curso_nombre, m.curso_categoria, m.jornada_descripcion,
            m.sede, m.get_modalidad_display(), m.jornada_horario,
            m.tipo_matricula, m.estado,
            m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '',
            float(m.valor_curso), float(m.descuento), float(m.valor_neto),
            float(m.valor_pagado), float(m.saldo), m.estado_pago,
            m.talla_camiseta, m.tipo_registro, m.factura_realizada,
            m.observaciones, m.registrado_por_nombre,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            # Columnas monetarias en negrita verde
            if col_idx in (18, 19, 20, 21, 22):
                cell.font = money_font
                cell.number_format = '"$"#,##0.00'

    # Totales
    if matriculas:
        total_row = len(matriculas) + 5
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = total_fill
        for col in range(1, 18):
            ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=18, value=float(sum(m.valor_curso for m in matriculas))).fill = total_fill
        ws.cell(row=total_row, column=19, value=float(sum(m.descuento for m in matriculas))).fill = total_fill
        ws.cell(row=total_row, column=20, value=float(sum(m.valor_neto for m in matriculas))).fill = total_fill
        ws.cell(row=total_row, column=21, value=float(sum(m.valor_pagado for m in matriculas))).fill = total_fill
        ws.cell(row=total_row, column=22, value=float(sum(m.saldo for m in matriculas))).fill = total_fill
        for col in range(18, 23):
            c = ws.cell(row=total_row, column=col)
            c.font = Font(bold=True, color='1A237E')
            c.number_format = '"$"#,##0.00'

    # Anchos
    anchos = [14, 18, 18, 6, 26, 14, 14, 14, 24, 14, 22, 14, 12, 14, 16, 14, 13, 12, 12, 12, 12, 12, 12, 6, 14, 8, 30, 18]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a

    # ── Hoja 2: Abonos ──
    ws2 = wb.create_sheet('Abonos')

    headers2 = [
        'Cédula estudiante', 'Estudiante', 'Curso', 'Jornada', 'Modalidad',
        'Fecha abono', 'Recibo', 'Monto', 'Tipo pago', 'Módulo',
        'Cuenta saldo', 'Método', 'Banco', 'Observaciones', 'Registrado por',
    ]

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers2))
    ws2.cell(row=1, column=1, value=f'Abonos archivados — {cierre.encabezado}').font = title_font
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 24

    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws2.row_dimensions[3].height = 30

    row = 4
    for m in matriculas:
        for a in m.abonos_archivados.all().order_by('fecha'):
            ws2.cell(row=row, column=1, value=m.cedula).border = thin_border
            ws2.cell(row=row, column=2, value=f'{m.apellidos} {m.nombres}').border = thin_border
            ws2.cell(row=row, column=3, value=m.curso_nombre).border = thin_border
            ws2.cell(row=row, column=4, value=m.jornada_descripcion).border = thin_border
            ws2.cell(row=row, column=5, value=m.get_modalidad_display()).border = thin_border
            ws2.cell(row=row, column=6, value=a.fecha.strftime('%d/%m/%Y') if a.fecha else '').border = thin_border
            ws2.cell(row=row, column=7, value=a.numero_recibo).border = thin_border
            c8 = ws2.cell(row=row, column=8, value=float(a.monto))
            c8.font = money_font
            c8.number_format = '"$"#,##0.00'
            c8.border = thin_border
            ws2.cell(row=row, column=9, value=a.tipo_pago_label).border = thin_border
            ws2.cell(row=row, column=10, value=a.numero_modulo or '').border = thin_border
            ws2.cell(row=row, column=11, value='Sí' if a.cuenta_para_saldo else 'No').border = thin_border
            ws2.cell(row=row, column=12, value=a.metodo_label).border = thin_border
            ws2.cell(row=row, column=13, value=a.banco_label).border = thin_border
            ws2.cell(row=row, column=14, value=a.observaciones).border = thin_border
            ws2.cell(row=row, column=15, value=a.registrado_por_nombre).border = thin_border
            row += 1

    anchos2 = [16, 26, 22, 22, 12, 13, 12, 12, 16, 8, 11, 16, 14, 30, 18]
    for i, a in enumerate(anchos2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = a

    # ── Respuesta ──
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    nombre_archivo = (
        f'cierre_{cierre.curso_nombre[:30]}_'
        f'{cierre.fecha_cierre.strftime("%Y%m%d")}'
    ).replace(' ', '_').replace('/', '-')

    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    return resp


# ═════════════════════════════════════════════════════════════════
# Estudiantes archivados (directorio histórico)
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def estudiantes_archivados_lista(request):
    """
    Lista todos los estudiantes archivados (los que se borraron del directorio
    durante un cierre con la opción 'Limpiar directorio'). Con búsqueda y filtros.
    """
    q = request.GET.get('q', '').strip()
    cierre_id = request.GET.get('cierre', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()

    qs = EstudianteArchivado.objects.select_related('cierre').order_by('apellidos', 'nombres')

    if q:
        qs = qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
            | Q(correo__icontains=q)
            | Q(celular__icontains=q)
        )
    if cierre_id.isdigit():
        qs = qs.filter(cierre_id=int(cierre_id))
    if anio.isdigit():
        qs = qs.filter(archivado_en__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(archivado_en__month=int(mes))

    cierres_disponibles = CierreCurso.objects.filter(
        estudiantes_archivados__isnull=False
    ).distinct().order_by('-fecha_cierre')

    anios = sorted(
        set(EstudianteArchivado.objects.dates('archivado_en', 'year').values_list('archivado_en__year', flat=True)),
        reverse=True
    )

    return render(request, 'estudiantes/archivados_lista.html', {
        'estudiantes': qs,
        'cierres_disponibles': cierres_disponibles,
        'anios': anios,
        'total': qs.count(),
        'filtros': {'q': q, 'cierre': cierre_id, 'anio': anio, 'mes': mes},
    })


@matricula_requerida
def estudiantes_archivados_export(request):
    """Descarga los estudiantes archivados como Excel (respetando filtros)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    q = request.GET.get('q', '').strip()
    cierre_id = request.GET.get('cierre', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()

    qs = EstudianteArchivado.objects.select_related('cierre').order_by('apellidos', 'nombres')
    if q:
        qs = qs.filter(
            Q(cedula__icontains=q) | Q(apellidos__icontains=q) | Q(nombres__icontains=q)
            | Q(correo__icontains=q) | Q(celular__icontains=q)
        )
    if cierre_id.isdigit():
        qs = qs.filter(cierre_id=int(cierre_id))
    if anio.isdigit():
        qs = qs.filter(archivado_en__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(archivado_en__month=int(mes))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Estudiantes archivados'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Edad', 'Correo', 'Celular',
        'Ciudad', 'Nivel formación', 'Título profesional',
        'Fecha registro original', 'Archivado el', 'Cierre asociado',
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value='Estudiantes archivados — Directorio histórico')
    title.font = Font(bold=True, size=14, color='1A237E')
    title.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
    ws.row_dimensions[3].height = 30

    for row_idx, e in enumerate(qs, start=4):
        vals = [
            e.cedula, e.apellidos, e.nombres, e.edad or '',
            e.correo, e.celular, e.ciudad, e.nivel_formacion, e.titulo_profesional,
            e.creado_original.strftime('%d/%m/%Y') if e.creado_original else '',
            e.archivado_en.strftime('%d/%m/%Y %H:%M') if e.archivado_en else '',
            e.cierre.encabezado if e.cierre else '',
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical='center')

    anchos = [16, 20, 20, 6, 26, 14, 14, 22, 24, 14, 18, 36]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="estudiantes_archivados_'
        f'{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    return resp